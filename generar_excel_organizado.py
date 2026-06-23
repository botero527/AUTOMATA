#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Generador de Excel organizados por vehículo y carpeta
Filtra solo archivos con datos y organiza por tipo de pieza
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Ruta del archivo JSON de entrada
INPUT_JSON = "output/SUZUKI_TODO.json"
OUTPUT_EXCEL = "output/planos_tecnicos_organizado.xlsx"

# Definición de tipos de pieza
TIPOS_PIEZA = {
    "000": "Parabrisas",
    "001": "Lateral izquierdo",
    "002": "Lateral derecho",
    "003": "Lateral trasero Izq",
    "004": "Lateral trasero Der",
    "005": "Ventilete trasero Izq",
    "006": "Ventilete trasero Der",
    "007": "Cabina izquierda",
    "008": "Cabina derecha",
    "009": "Posterior/Luneta",
    "010": "Sunroof principal",
    "011": "Lateral extendido Izq",
    "012": "Lateral extendido Der",
    "019": "Ventilete adelante Izq",
    "020": "Ventilete adelante Der",
    "025": "Sunroof secundario",
    "087": "Sunroof terciario",
    "090": "Sunroof panorámico"
}

def get_tipo_pieza(nombre_archivo):
    """Extraer código de tipo de pieza del nombre del archivo"""
    import re
    nombre = nombre_archivo.replace(".dwg", "").replace(".DWG", "")
    match = re.search(r'(\d{2,3})\s', nombre)
    if match:
        codigo = match.group(1)
        if len(codigo) == 2:
            codigo = '0' + codigo
        return codigo
    return None

def get_tipo_nombre(codigo):
    """Obtener nombre del tipo de pieza"""
    return TIPOS_PIEZA.get(codigo, "Otro")

def cargar_datos():
    """Cargar datos del JSON"""
    with open(INPUT_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)

def crear_excel_organizado(datos):
    """Crear Excel con estructura optimizada para comparación"""
    wb = Workbook()
    wb.remove(wb.active)  # Quitar hoja por defecto
    
    # Estilos
    header_fill = PatternFill(start_color="00A2D7", end_color="00A2D7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="00A2D7")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agrupar planos por carpeta
    carpetas = defaultdict(list)
    for plano in datos['planos']:
        if plano.get('tablas') and len(plano.get('tablas', [])) > 0:
            carpeta = plano.get('carpeta', 'Sin carpeta')
            carpetas[carpeta].append(plano)
    
    # Crear una hoja por carpeta
    for carpeta, planos in sorted(carpetas.items()):
        ws = wb.create_sheet(title=carpeta[:31])  # Excel limita nombres a 31 chars
        
        # Encabezados
        headers = ['Archivo', 'Tipo Pieza', 'Nombre Tipo']
        
        # Agregar todos los campos únicos como columnas
        campos_global = set()
        for plano in planos:
            for tabla in plano.get('tablas', []):
                for campo in tabla.get('atributos', {}).keys():
                    campos_global.add(campo)
        
        headers.extend(sorted(campos_global))
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Escribir datos
        for row, plano in enumerate(planos, 2):
            # Archivo y tipo
            ws.cell(row=row, column=1, value=plano['archivo'])
            
            tipo = get_tipo_pieza(plano['archivo'])
            ws.cell(row=row, column=2, value=tipo or '')
            ws.cell(row=row, column=3, value=get_tipo_nombre(tipo) if tipo else '')
            
            # Campos de las tablas
            col_idx = 4
            for campo in sorted(campos_global):
                valor = ''
                for tabla in plano.get('tablas', []):
                    if campo in tabla.get('atributos', {}):
                        valor = tabla['atributos'][campo]
                        break
                cell = ws.cell(row=row, column=col_idx, value=valor)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                col_idx += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30  # Archivo
        ws.column_dimensions['B'].width = 12  # Tipo
        ws.column_dimensions['C'].width = 25  # Nombre tipo
        for i, campo in enumerate(sorted(campos_global), 4):
            ws.column_dimensions[get_column_letter(i)].width = 20
    
    # Crear hoja RESUMEN
    ws_resumen = wb.create_sheet(title="RESUMEN", index=0)
    
    # Encabezados del resumen
    res_headers = ['Carpeta', 'Total Archivos', 'Con Datos', 'Por Tipo de Pieza']
    for col, header in enumerate(res_headers, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Datos del resumen
    row = 2
    for carpeta, planos in sorted(carpetas.items()):
        total = len(planos)
        
        # Contar por tipo de pieza
        tipos_count = defaultdict(int)
        for plano in planos:
            tipo = get_tipo_pieza(plano['archivo'])
            if tipo:
                tipos_count[tipo] += 1
        
        tipos_str = ', '.join([f"{get_tipo_nombre(k)}: {v}" for k, v in sorted(tipos_count.items())])
        
        ws_resumen.cell(row=row, column=1, value=carpeta)
        ws_resumen.cell(row=row, column=2, value=total)
        ws_resumen.cell(row=row, column=3, value=total)  # Todos tienen datos (ya filtramos)
        ws_resumen.cell(row=row, column=4, value=tipos_str)
        
        for col in range(1, 5):
            ws_resumen.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 60
    
    # Guardar
    wb.save(OUTPUT_EXCEL)
    print(f"Excel creado: {OUTPUT_EXCEL}")
    print(f"   - Hoja RESUMEN con estadisticas por carpeta")
    print(f"   - {len(carpetas)} hojas de carpetas con datos")

if __name__ == "__main__":
    datos = cargar_datos()
    crear_excel_organizado(datos)