"""
EXTRACTOR JEEP - Similar a extraer_090_toyota.py
================================================
- Convierte archivo por archivo usando ODA
- Extrae campos técnicos (OFFSET, ACERO, BN+D, etc.)
- Busca imágenes en la base de datos
"""
import os
import sys
import json
import subprocess
import shutil
import tempfile
import time
import re
import pyodbc
from datetime import datetime

# ============== CONFIGURACIÓN ==============
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
VEHICULO = "SUZUKI"

ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

# ============== CONFIGURACIÓN BD ==============
SERVER_BD = 'agpcolsap.database.windows.net'
DB_BD = 'DB_COL_SAP'
USER_BD = 'Viewer'
PSW_BD = 'AgpconsCol2023'

# Campos técnicos a buscar
CAMPOS_A_BUSCAR = [
    "OFFSET", "BN+D", "BN INT", "ACERO", "STEEL", "ESPESOR", 
    "LARGO", "ANCHO", "PESO", "MATERIAL", "BANDA", "TIPO", "MEDIDA"
]

import ezdxf


def buscar_imagenes_en_bd(nombre_dwg):
    """
    Busca imágenes en la base de datos SAP usando el número de documento
    """
    imagenes = []
    try:
        nombre_base = os.path.splitext(nombre_dwg)[0].upper()
        numeros = re.split(r'[^0-9]+', nombre_base)
        numeros = [n for n in numeros if n]
        
        if not numeros:
            return []
        
        if len(numeros) >= 3:
            codigo = f"M{numeros[0]} {numeros[1]} {numeros[2]}"
        elif len(numeros) == 2:
            codigo = f"M{numeros[0]} {numeros[1]}"
        else:
            codigo = f"M{numeros[0]}"
        
        conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={SERVER_BD};DATABASE={DB_BD};UID={USER_BD};PWD={PSW_BD}'
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        query = """
            SELECT PLANO 
            FROM [dbo].[ODATA_ZFER_RUTAS_JPG] 
            WHERE PLANO LIKE ? OR DOCUMENTO LIKE ?
        """
        
        param = f"%{codigo}%"
        cursor.execute(query, (param, param))
        
        for row in cursor.fetchall():
            if row.PLANO:
                imagenes.append(row.PLANO)
        
        conn.close()
        
        if imagenes:
            print_progress(f"    -> BD: {len(imagenes)} imágenes", "OK")
            
    except Exception as e:
        print_progress(f"    -> Error BD: {e}", "ERR")
    
    return imagenes[:10]


def print_progress(msg, tipo="INFO"):
    hora = datetime.now().strftime("%H:%M:%S")
    print(f"[{hora}] [{tipo}] {msg}")
    sys.stdout.flush()


def buscar_archivos(servidor, vehiculo):
    """Busca TODOS los archivos DWG del vehículo"""
    print_progress(f"Buscando archivos DWG en {vehiculo}...")
    ruta_vehiculo = os.path.join(servidor, vehiculo)
    archivos = []
    
    for root, dirs, files in os.walk(ruta_vehiculo):
        if '_DXF' in root or '_dxf' in root:
            continue
        for f in files:
            if f.lower().endswith('.dwg'):
                archivos.append({
                    'ruta': os.path.join(root, f),
                    'nombre': f,
                    'carpeta': os.path.basename(root)
                })
    
    print_progress(f"Encontrados {len(archivos)} archivos DWG", "OK")
    return archivos


def convertir_dwg_a_dxf(dwg_path, temp_dir):
    """Convierte un DWG a DXF usando ODA (igual que antes)"""
    try:
        carpeta = os.path.dirname(dwg_path)
        nombre = os.path.basename(dwg_path)
        
        # ODA: input_folder output_folder version format recurse silent
        cmd = [ODA_EXE, carpeta, temp_dir, "ACAD2018", "DXF", "0", "0"]
        subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        dxf_path = os.path.join(temp_dir, os.path.splitext(nombre)[0] + ".dxf")
        return dxf_path if os.path.exists(dxf_path) else None
    except Exception as e:
        print_progress(f"Error converting: {e}", "ERR")
        return None


def extraer_campos_tecnicos(dxf_path):
    """Busca campos técnicos específicos en el DXF"""
    try:
        doc = ezdxf.readfile(dxf_path)
        msp = doc.modelspace()
        
        textos = []
        
        # TEXT
        for text in msp.query('TEXT'):
            try:
                x = text.dxf.insert.x
                y = text.dxf.insert.y
                txt = text.dxf.text.strip()
                if txt:
                    textos.append({'x': x, 'y': y, 'texto': txt})
            except:
                pass
        
        # MTEXT
        for mtext in msp.query('MTEXT'):
            try:
                if hasattr(mtext.dxf, 'insert'):
                    x = mtext.dxf.insert.x
                    y = mtext.dxf.insert.y
                else:
                    x, y = 0, 0
                txt = mtext.text.strip() if hasattr(mtext, 'text') else ''
                if txt:
                    txt = txt.replace('\\P', ' ').replace('\\~', ' ').strip()
                    textos.append({'x': x, 'y': y, 'texto': txt})
            except:
                pass
        
        # ATTRIB
        for block_ref in msp.query('INSERT'):
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                try:
                    bx = block_ref.dxf.insert.x if hasattr(block_ref.dxf, 'insert') else 0
                    by = block_ref.dxf.insert.y if hasattr(block_ref.dxf, 'insert') else 0
                except:
                    bx, by = 0, 0
                
                for attr in block_ref.attribs:
                    try:
                        tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                        txt = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                        if tag and txt:
                            textos.append({'x': bx, 'y': by, 'texto': tag})
                            textos.append({'x': bx + 10, 'y': by, 'texto': txt})
                    except:
                        pass
        
        if not textos:
            return []
        
        # BUSCAR PALABRAS CLAVE Y SUS VALORES
        datos_encontrados = {}
        
        for t in textos:
            texto_upper = t['texto'].upper().strip()
            
            es_campo = False
            for kw in CAMPOS_A_BUSCAR:
                if kw.upper() in texto_upper or texto_upper.startswith(kw.upper()):
                    es_campo = True
                    break
            
            if es_campo:
                campo = t['texto'].strip()
                valores_encontrados = []
                
                # Valores a la DERECHA
                for otro in textos:
                    if abs(otro['y'] - t['y']) < 2 and otro['x'] > t['x']:
                        distancia = otro['x'] - t['x']
                        if distancia < 80:
                            otro_upper = otro['texto'].upper()
                            es_otra_kw = any(kw.upper() in otro_upper or otro_upper.startswith(kw.upper()) for kw in CAMPOS_A_BUSCAR)
                            if not es_otra_kw and otro['texto'].strip():
                                valores_encontrados.append(otro['texto'].strip())
                
                # Valores ABAJO
                textos_debajo = [otro for otro in textos if otro['y'] < t['y'] - 2 and otro['y'] > t['y'] - 30]
                textos_debajo.sort(key=lambda x: x['y'], reverse=True)
                
                for otro in textos_debajo:
                    if abs(otro['x'] - t['x']) < 10:
                        otro_upper = otro['texto'].upper()
                        es_otra_kw = any(kw.upper() in otro_upper or otro_upper.startswith(kw.upper()) for kw in CAMPOS_A_BUSCAR)
                        if not es_otra_kw and otro['texto'].strip():
                            valores_encontrados.append(otro['texto'].strip())
                
                if valores_encontrados:
                    valor_final = " | ".join(valores_encontrados)
                    if campo in datos_encontrados:
                        if valor_final not in datos_encontrados[campo]:
                            datos_encontrados[campo] = datos_encontrados[campo] + " | " + valor_final
                    else:
                        datos_encontrados[campo] = valor_final
        
        if datos_encontrados:
            return [{'atributos': datos_encontrados}]
        
        return []
        
    except Exception as e:
        print_progress(f"Error: {e}", "ERR")
        return []


def procesar_archivos(archivos, temp_dir):
    """Procesa cada archivo"""
    resultados = []
    exitosos = 0
    errores = 0
    
    total = len(archivos)
    
    for i, info in enumerate(archivos, 1):
        print_progress(f"[{i}/{total}] {info['nombre']}", "PROC")
        
        # Buscar imágenes en la base de datos
        imagenes = buscar_imagenes_en_bd(info['nombre'])
        
        dxf_path = convertir_dwg_a_dxf(info['ruta'], temp_dir)
        
        if dxf_path and os.path.exists(dxf_path):
            tablas = extraer_campos_tecnicos(dxf_path)
            
            if tablas and tablas[0].get('atributos'):
                exitosos += 1
                resumen = tablas[0]['atributos']
                
                for campo, valor in list(resumen.items())[:3]:
                    print_progress(f"  -> {campo}: {valor}", "DAT")
                
                resultados.append({
                    'archivo': info['nombre'],
                    'carpeta': info['carpeta'],
                    'tablas': tablas,
                    'resumen': resumen,
                    'imagenes': imagenes
                })
            else:
                print_progress(f"  -> Sin datos técnicos", "WARN")
                errores += 1
                resultados.append({
                    'archivo': info['nombre'],
                    'carpeta': info['carpeta'],
                    'tablas': [],
                    'resumen': {},
                    'imagenes': imagenes
                })
        else:
            print_progress(f"  -> Sin convertir", "WARN")
            errores += 1
            resultados.append({
                'archivo': info['nombre'],
                'carpeta': info['carpeta'],
                'tablas': [],
                'resumen': {},
                'imagenes': imagenes
            })
    
    return resultados, exitosos, errores


def guardar_json(resultados, exitosos, errores, output_path):
    datos = {
        "metadata": {
            "vehiculo": VEHICULO,
            "total_archivos": len(resultados),
            "exitosos": exitosos,
            "errores": errores,
            "fecha_extraccion": datetime.now().isoformat()
        },
        "planos": resultados
    }
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)
    
    print_progress(f"JSON: {output_path}", "OK")


def guardar_excel(resultados, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planos"
        
        headers = ["Archivo", "Carpeta", "Campo", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="00A2D7", end_color="00A2D7", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row = 2
        for plano in resultados:
            archivo = plano['archivo']
            carpeta = plano['carpeta']
            
            if plano['tablas']:
                for t in plano['tablas']:
                    for campo, valor in t.get('atributos', {}).items():
                        ws.cell(row=row, column=1, value=archivo)
                        ws.cell(row=row, column=2, value=carpeta)
                        ws.cell(row=row, column=3, value=campo)
                        ws.cell(row=row, column=4, value=valor)
                        row += 1
            else:
                ws.cell(row=row, column=1, value=archivo)
                ws.cell(row=row, column=2, value=carpeta)
                ws.cell(row=row, column=3, value="(sin datos)")
                ws.cell(row=row, column=4, value="")
                row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print_progress(f"Excel: {output_path}", "OK")
        
    except Exception as e:
        print_progress(f"Error Excel: {e}", "ERR")


def main():
    print("=" * 60)
    print(f"  EXTRACTOR {VEHICULO} - Todos los archivos")
    print("=" * 60)
    print()
    
    archivos = buscar_archivos(SERVIDOR, VEHICULO)
    
    if not archivos:
        print("No se encontraron archivos")
        return
    
    print()
    print_progress(f"Procesando {len(archivos)} archivos...")
    print()
    
    with tempfile.TemporaryDirectory() as temp_dir:
        resultados, exitosos, errores = procesar_archivos(archivos, temp_dir)
    
    print()
    print("=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    print(f"  Total:       {len(archivos)}")
    print(f"  Exitosos:   {exitosos}")
    print(f"  Errores:    {errores}")
    print(f"  JSON:       output/{VEHICULO}_TODO.json")
    print(f"  Excel:      output/{VEHICULO}_TODO.xlsx")
    print("=" * 60)
    
    guardar_json(resultados, exitosos, errores, f"output/{VEHICULO}_TODO.json")
    guardar_excel(resultados, f"output/{VEHICULO}_TODO.xlsx")


if __name__ == "__main__":
    main()